"""
AWP 5.2 & 5.8 Automation
Royal Audit Authority - Supreme Audit Institution of Bhutan

This module is organized in two halves:
  1. Pure, side-effect-free business logic (text cleaning, PDF table
     extraction/stitching, workbook generation). None of these functions
     touch Streamlit or global state, so they are cheap to unit test and
     safe to wrap in @st.cache_data.
  2. Streamlit UI/orchestration code, guarded by `if __name__ == "__main__"`
     so this file can be imported (e.g. by test_extraction.py) without
     starting a Streamlit session.
"""

import os
import re
import math
import time
import hashlib
import logging
import unicodedata
from io import BytesIO
from dataclasses import dataclass

import streamlit as st
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ===================================================================
# Logging
# ===================================================================
logger = logging.getLogger("awp_automation")
if not logger.handlers:
    handler = logging.StreamHandler()
    handler.setFormatter(logging.Formatter(
        "%(asctime)s [%(levelname)s] %(name)s: %(message)s"
    ))
    logger.addHandler(handler)
logger.setLevel(logging.INFO)

DEBUG_EXTRACTION = os.getenv("DEBUG_EXTRACTION", "false").lower() == "true"

# ===================================================================
# Safeguards / limits
# ===================================================================
MAX_PDF_SIZE_BYTES = 40 * 1024 * 1024   # 40 MB
MAX_PDF_PAGES = 150


class PDFValidationError(Exception):
    """Raised for any PDF that fails safety/sanity checks before parsing."""


# ===================================================================
# Text normalization
#
# clean_extracted_text()  -> for DATA (never truncated, never restricted
#                             to Excel's 31-char worksheet-name rule)
# sanitize_sheet_name()   -> for worksheet TAB names only
# ===================================================================
_ILLEGAL_XML_CHARS = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]')
_MULTI_WS = re.compile(r'\s+')
_DASH_VARIANTS = re.compile(r'[\u2010\u2011\u2012\u2013\u2014\u2015]')


def normalize_ws(s: str) -> str:
    """Collapse any run of whitespace (including line breaks/tabs) to a
    single space, and strip non-breaking spaces to regular spaces."""
    if not s:
        return ""
    s = s.replace("\u00a0", " ").replace("\t", " ")
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    return _MULTI_WS.sub(" ", s).strip()


def clean_extracted_text(value) -> str:
    """Full-fidelity cleaner for extracted PDF text / business data.

    - Converts None -> ""
    - Removes illegal XML/control characters (openpyxl would reject them)
    - Joins wrapped lines with a single space
    - Normalizes repeated whitespace and non-breaking spaces
    - Normalizes unicode dash variants to a standard en dash for consistency
    - Preserves meaningful punctuation: & , / - EN-DASH ( ) . ; : etc.
    - NEVER truncates. NEVER applies the 31-character Excel worksheet limit.
    """
    if value is None:
        return ""
    s = str(value)
    s = _ILLEGAL_XML_CHARS.sub('', s)
    s = unicodedata.normalize("NFKC", s)
    s = _DASH_VARIANTS.sub('\u2013', s)  # normalize all dash variants -> en dash
    s = normalize_ws(s)
    return s


_SHEET_ILLEGAL_CHARS = re.compile(r'[\\/*?:\[\]]')


def sanitize_sheet_name(value: str) -> str:
    """Remove ONLY the characters Excel forbids in a worksheet tab name.
    Does not touch length - callers apply the 31-char limit separately.
    Also strips leading/trailing apostrophes (Excel disallows tabs that
    start or end with a single quote) and guards against an empty name.
    """
    name = clean_extracted_text(value)
    name = _SHEET_ILLEGAL_CHARS.sub('', name)
    name = name.strip().strip("'").strip()
    return name or "Sheet"


def make_excel_sheet_name(value: str, used_names: dict) -> str:
    """Produce a unique, <=31-char, Excel-legal worksheet tab name from a
    (potentially much longer) full class/COTABD name. `used_names` is a
    dict the caller keeps across calls in a single workbook, tracking how
    many times a given base name has been used so collisions get _2, _3, ...
    suffixes without ever exceeding 31 characters.
    """
    base_full = sanitize_sheet_name(value)
    base = base_full[:31]
    if base not in used_names:
        used_names[base] = 0
        return base
    used_names[base] += 1
    suffix = f"_{used_names[base]}"
    trimmed = base[:max(1, 31 - len(suffix))]
    candidate = trimmed + suffix
    while candidate in used_names:
        used_names[base] += 1
        suffix = f"_{used_names[base]}"
        trimmed = base[:max(1, 31 - len(suffix))]
        candidate = trimmed + suffix
    used_names[candidate] = 0
    return candidate


# Backwards-compatible alias kept in case other code imports the old name.
def clean_title(t):
    return sanitize_sheet_name(t)[:31]


def clean_cell(v):
    """Legacy name kept for compatibility; delegates to clean_extracted_text."""
    return clean_extracted_text(v)


_CONTINUATION_TRAILERS = (",", "&", "-", "\u2013", "\u2014", "(", ":", ";", "/")
_CONTINUATION_WORDS = {
    "and", "or", "the", "of", "to", "for", "with", "a", "an", "in", "on",
    "&", "-",
}


def looks_incomplete(text: str) -> bool:
    """Heuristic: does this text look like it was cut off mid-thought by a
    PDF line wrap (and therefore likely continues on the next physical
    row)? Used to decide whether to merge adjacent extracted rows instead
    of blindly concatenating every row, or blindly stopping at every row.
    """
    if not text:
        return False
    t = text.strip()
    if t.endswith(_CONTINUATION_TRAILERS):
        return True
    words = t.split()
    if words and words[-1].strip(".,;:").lower() in _CONTINUATION_WORDS:
        return True
    return False


def split_assertions(text):
    if not text:
        return []
    parts = re.split(r'\s+and\s+|\s*&\s*|\s*/\s*|,\s*|;\s*', text, flags=re.IGNORECASE)
    return [p.strip() for p in parts if p.strip()]


def normalize_header_text(s: str) -> str:
    """Lowercase, whitespace-collapsed text used purely for keyword
    matching against table headers. Never used for display."""
    return normalize_ws(clean_extracted_text(s)).lower()


# ===================================================================
# PDF -> raw per-page table extraction (single read of the PDF)
# ===================================================================
@dataclass
class ExtractionResult:
    tables_by_page: list          # [(page_num:int, [table, table, ...]), ...] table = list[list[str]]
    header_info: dict
    num_pages: int
    num_tables: int
    pdf_hash_prefix: str


def compute_pdf_hash(pdf_bytes: bytes) -> str:
    return hashlib.sha256(pdf_bytes).hexdigest()


def validate_pdf_bytes(pdf_bytes: bytes, max_size=MAX_PDF_SIZE_BYTES, max_pages=MAX_PDF_PAGES):
    """Cheap, fast sanity checks before we spend CPU time parsing tables."""
    if not pdf_bytes:
        raise PDFValidationError("The uploaded file is empty.")
    if len(pdf_bytes) > max_size:
        raise PDFValidationError(
            f"The PDF is too large ({len(pdf_bytes) / (1024*1024):.1f} MB). "
            f"Maximum allowed size is {max_size / (1024*1024):.0f} MB."
        )
    if pdf_bytes[:4] != b"%PDF":
        raise PDFValidationError("This file does not look like a valid PDF.")
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            n_pages = len(pdf.pages)
    except Exception as exc:
        msg = str(exc).lower()
        if "password" in msg or "encrypt" in msg:
            raise PDFValidationError(
                "This PDF appears to be encrypted/password-protected and cannot be read."
            )
        raise PDFValidationError(f"This PDF could not be opened ({exc.__class__.__name__}).")
    if n_pages == 0:
        raise PDFValidationError("The PDF has no pages.")
    if n_pages > max_pages:
        raise PDFValidationError(
            f"The PDF has {n_pages} pages, which exceeds the maximum of {max_pages}."
        )
    return n_pages


def _raw_table_to_str_rows(table):
    rows = []
    for row in table:
        rows.append(["" if c is None else str(c) for c in row])
    return rows


def extract_pdf_tables(pdf_bytes: bytes) -> ExtractionResult:
    """Read the PDF exactly once: pull every page's tables (in page
    order) plus the header/signature block info. Pure function - no
    Streamlit calls - so it is safe to wrap with @st.cache_data."""
    n_pages = validate_pdf_bytes(pdf_bytes)
    pdf_hash_prefix = compute_pdf_hash(pdf_bytes)[:12]

    tables_by_page = []
    num_tables = 0
    has_any_text = False

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            page_tables_raw = page.extract_tables() or []
            page_tables = [_raw_table_to_str_rows(t) for t in page_tables_raw if t]
            num_tables += len(page_tables)
            tables_by_page.append((page_num, page_tables))
            if not has_any_text:
                txt = page.extract_text() or ""
                if txt.strip():
                    has_any_text = True

    if not has_any_text and num_tables == 0:
        raise PDFValidationError(
            "No extractable text or tables were found. This PDF may be a scanned "
            "image - OCR would be required, which this tool does not perform automatically."
        )

    all_tables_flat = [t for _, tables in tables_by_page for t in tables]
    header_info = extract_header_info(all_tables_flat)

    logger.info(
        "pdf_hash=%s pages=%d tables=%d header_extracted=%s",
        pdf_hash_prefix, n_pages, num_tables,
        bool(header_info.get("entity_name")),
    )

    return ExtractionResult(
        tables_by_page=tables_by_page,
        header_info=header_info,
        num_pages=n_pages,
        num_tables=num_tables,
        pdf_hash_prefix=pdf_hash_prefix,
    )


def extract_header_info(all_tables):
    info = {k: "" for k in ["entity_name", "audit_period", "assessed_name", "assessed_designation",
                             "assessed_date", "reviewed_name", "reviewed_designation", "reviewed_date"]}
    for table in all_tables:
        if not table:
            continue
        for ri, row in enumerate(table):
            rt = [clean_extracted_text(c) for c in row]
            for i, cell in enumerate(rt):
                lc = cell.lower()

                def next_val(col_offset=1):
                    vals = [x for x in rt[i + col_offset:] if x]
                    if vals:
                        return vals[0]
                    if ri + 1 < len(table):
                        nxt = [clean_extracted_text(c) for c in table[ri + 1]]
                        v = [x for x in nxt if x]
                        if v:
                            return v[0]
                    return ""

                if "name of the entity" in lc:
                    info["entity_name"] = next_val() or re.sub(
                        r'name of the entity', '', cell, flags=re.IGNORECASE).strip(" :-")
                elif "period of audit" in lc:
                    m = re.search(r'(\d{4}[-/]\d{2,4})', " ".join(rt))
                    info["audit_period"] = next_val() or (m.group(1) if m else "")
                elif lc in ("name:", "name"):
                    vals = [x for x in rt[i + 1:] if x]
                    if len(vals) >= 1:
                        info["assessed_name"] = vals[0]
                    if len(vals) >= 2:
                        info["reviewed_name"] = vals[1]
                elif lc in ("designation:", "designation"):
                    vals = [x for x in rt[i + 1:] if x]
                    if len(vals) >= 1:
                        info["assessed_designation"] = vals[0]
                    if len(vals) >= 2:
                        info["reviewed_designation"] = vals[1]
                elif lc in ("date:", "date"):
                    vals = [x for x in rt[i + 1:] if x]
                    if len(vals) >= 1:
                        info["assessed_date"] = vals[0]
                    if len(vals) >= 2:
                        info["reviewed_date"] = vals[1]
    return info


# ===================================================================
# Table C ("Not Significant / NOT MATERIAL COTABD") detection and
# multi-page stitching.
#
# Table C is identified structurally (not by hardcoded class names): it
# is the table whose header row labels a "not significant / not
# material" column next to a "minimum audit procedure" column. Once
# found, we keep collecting rows from subsequent tables/pages until we
# hit clear evidence of a new, unrelated section (a wide AWP 5.2-style
# table, or document metadata like "Prepared by" / "STEP 1").
# ===================================================================
_NOT_SIGNIFICANT_MARKERS = (
    "not significant", "non significant", "non-significant", "not material",
)
_MAP_PROCEDURE_MARKERS = (
    "minimum audit", "analytical procedure", "audit procedure",
)
_SECTION_METADATA_MARKERS = (
    "name of the entity", "period of audit", "prepared by", "reviewed",
    "step 1", "step 2", "overall conclusion", "designation:", "date:",
    "signature",
)


def is_table_c_header_row(row) -> bool:
    """Flexible, normalized match for the Table C header - tolerates the
    header being split across extra empty columns."""
    cells = [normalize_header_text(c) for c in row]
    non_empty = [c for c in cells if c]
    if len(non_empty) < 2:
        return False
    joined = " ".join(non_empty)
    has_not_sig = any(m in joined for m in _NOT_SIGNIFICANT_MARKERS)
    has_map = any(m in joined for m in _MAP_PROCEDURE_MARKERS)
    return has_not_sig and has_map


def detect_table_c_start(table):
    """Search every row of a table (not just row 0) for the Table C
    header, since it may appear after blank rows or a title row."""
    for idx, row in enumerate(table):
        if is_table_c_header_row(row):
            return idx
    return None


def is_row_blank(row) -> bool:
    return all(not clean_extracted_text(c) for c in row)


def looks_like_new_major_section(row) -> bool:
    """Heuristic used to decide when to STOP collecting Table C
    continuation rows: either a wide (AWP 5.2-style) table row, or a row
    that is clearly document metadata / a new step heading."""
    cells = [clean_extracted_text(c) for c in row]
    non_empty = [c for c in cells if c]
    if len(non_empty) >= 5:
        return True
    joined = " ".join(c.lower() for c in non_empty)
    return any(m in joined for m in _SECTION_METADATA_MARKERS)


def flatten_row_to_two_columns(row):
    """Reduce an arbitrary-width extracted row down to a (name, proc)
    pair: first non-empty cell is the name candidate, everything else
    non-empty is joined (in order) as the procedure candidate. Handles
    PDFs that produce extra empty/split columns for what is logically a
    2-column table."""
    cells = [clean_extracted_text(c) for c in row]
    non_empty_idx = [i for i, c in enumerate(cells) if c]
    if not non_empty_idx:
        return "", ""
    first_idx = non_empty_idx[0]
    name = cells[first_idx]
    rest = [cells[i] for i in non_empty_idx if i != first_idx]
    proc = " ".join(rest)
    return name, proc


def stitch_table_c_segments(tables_by_page):
    """Walk pages/tables in document order, find where Table C starts,
    and collect every subsequent row that plausibly belongs to it -
    including continuation tables on later pages that do not repeat the
    header. Returns a flat list of raw (name, proc) row pairs (NOT yet
    merged across wrapped/split physical rows) plus the label text taken
    from the PDF's own header cell.
    """
    raw_rows = []
    label = ""
    started = False
    finished = False

    for page_num, tables in tables_by_page:
        if finished:
            break
        for table in tables:
            if finished:
                break
            if not table:
                continue

            if not started:
                start_idx = detect_table_c_start(table)
                if start_idx is None:
                    continue
                started = True
                header_row = table[start_idx]
                non_empty = [clean_extracted_text(c) for c in header_row if clean_extracted_text(c)]
                if non_empty:
                    label = non_empty[0].rstrip(" *:-")
                for row in table[start_idx + 1:]:
                    if is_row_blank(row):
                        continue
                    if is_table_c_header_row(row):
                        continue  # repeated header - not data
                    if looks_like_new_major_section(row):
                        finished = True
                        break
                    raw_rows.append(flatten_row_to_two_columns(row))
            else:
                # Already inside Table C: is this table a genuine
                # continuation, or a clearly different/new section?
                if not table:
                    continue
                first_meaningful = next((r for r in table if not is_row_blank(r)), None)
                if first_meaningful is not None and looks_like_new_major_section(first_meaningful) \
                        and not is_table_c_header_row(first_meaningful):
                    finished = True
                    break
                for row in table:
                    if is_row_blank(row):
                        continue
                    if is_table_c_header_row(row):
                        continue
                    if looks_like_new_major_section(row):
                        finished = True
                        break
                    raw_rows.append(flatten_row_to_two_columns(row))

    return raw_rows, (label or "Not Significant COTABD")


def merge_logical_rows(raw_rows):
    """Merge physically-split rows (wrapped class names / procedures,
    empty-cell continuations) into complete logical (name, procedure)
    records. Uses trailing-punctuation / dangling-conjunction evidence
    (`looks_incomplete`) rather than naively concatenating every row.
    """
    records = []
    i, n = 0, len(raw_rows)
    while i < n:
        name, proc = raw_rows[i]
        name, proc = clean_extracted_text(name), clean_extracted_text(proc)
        if not name and not proc:
            i += 1
            continue
        j = i + 1
        while j < n:
            next_name, next_proc = raw_rows[j]
            next_name = clean_extracted_text(next_name)
            next_proc = clean_extracted_text(next_proc)
            if not next_name and not next_proc:
                j += 1
                continue
            if not next_name and next_proc:
                # empty name cell -> continuation of procedure text
                proc = (proc + " " + next_proc).strip() if proc else next_proc
                j += 1
                continue
            if next_name and not next_proc and looks_incomplete(name):
                name = (name + " " + next_name).strip()
                j += 1
                continue
            if next_name and next_proc and (looks_incomplete(name) or looks_incomplete(proc)):
                name = (name + " " + next_name).strip()
                proc = (proc + " " + next_proc).strip()
                j += 1
                continue
            break
        records.append((name.strip(), proc.strip()))
        i = j
    return [(nm, pr) for nm, pr in records if nm and pr]


def extract_table_c_rows(tables_by_page):
    """Full Table C pipeline: stitch across pages -> merge split rows.
    Returns (rows, label) where rows is a de-duplicated list of
    (class_name, procedure_text) tuples with complete, untruncated text.
    """
    raw_rows, label = stitch_table_c_segments(tables_by_page)
    records = merge_logical_rows(raw_rows)

    seen = set()
    deduped = []
    for name, proc in records:
        key = (name.lower(), proc.lower())
        if key in seen:
            continue
        seen.add(key)
        deduped.append((name, proc))

    if DEBUG_EXTRACTION:
        logger.info("table_c_raw_rows=%d table_c_final_rows=%d label=%r",
                     len(raw_rows), len(deduped), label)
    return deduped, label


# Legacy-signature wrappers kept for anything that might import the old names.
def find_table_c(all_tables):
    for table in all_tables:
        idx = detect_table_c_start(table)
        if idx is not None:
            return table
    return None


def get_table_c_label(table):
    if not table:
        return "Not Significant COTABD"
    idx = detect_table_c_start(table)
    if idx is None:
        return "Not Significant COTABD"
    non_empty = [clean_extracted_text(c) for c in table[idx] if clean_extracted_text(c)]
    return (non_empty[0].rstrip(" *:-") if non_empty else "") or "Not Significant COTABD"


# ===================================================================
# AWP 5.2 (significant COTABD) parsing - with page/table continuation.
#
# A "class" (current_class) persists across tables and pages, exactly
# like the original implementation, but with:
#   - metadata rows (Prepared by / Reviewed / Date / Period of Audit /
#     page numbers) explicitly filtered out instead of relying on a
#     short hardcoded JUNK set;
#   - wrapped class-name continuation across rows/pages;
#   - continuation of a risk/control/substantive-procedure fragment
#     into the previous record when a row is otherwise empty and the
#     previous field looks cut off.
# ===================================================================
_SAP_JUNK_MARKERS = (
    "material", "risks", "name:", "date:", "designation:", "name of the entity",
    "prepared by", "reviewed", "reviewed & agreed by", "period of audit",
    "awp", "not significant", "signature", "step 1", "step 2",
)


_SAP_HEADER_KEYWORDS = ("cotabd", "risk", "assertion", "control activity", "substantive")


def is_sap_header_row(row) -> bool:
    """Detects a repeated AWP 5.2 column-header row (e.g. on a later
    page) via normalized keyword matching against individual, short
    cells - NOT a substring search over the whole row - so that a long
    data sentence which happens to contain a word like "risk" is never
    mistaken for a header."""
    cells = [normalize_header_text(c) for c in row]
    hits = 0
    for c in cells:
        if not c or len(c) > 40:
            continue
        for kw in _SAP_HEADER_KEYWORDS:
            if c == kw or (kw in c and len(c) <= len(kw) + 15):
                hits += 1
                break
    return hits >= 3


def is_metadata_row(row) -> bool:
    cells = [clean_extracted_text(c) for c in row]
    non_empty = [c for c in cells if c]
    if not non_empty:
        return True
    joined = " ".join(c.lower() for c in non_empty)
    if joined.strip().isdigit():
        return True
    return any(m in joined for m in _SAP_JUNK_MARKERS)


def _is_junk_class_candidate(raw: str) -> bool:
    lc = raw.lower()
    if not raw:
        return True
    if lc in _SAP_JUNK_MARKERS:
        return True
    return any(lc.startswith(j) for j in _SAP_JUNK_MARKERS)


def parse_sap_records(tables_by_page):
    """Returns an ordered dict: full_class_name -> list of record dicts
    with keys Risk / Assertions / Control Activity / Substantive Testing
    Procedures. Class names and text are never truncated."""
    parsed = {}
    current_class = None
    class_name_open = False  # True if current_class's own wrap continuation is still expected

    for page_num, tables in tables_by_page:
        relevant_tables = [t for t in tables if t and len(t[0]) >= 5]
        for table in relevant_tables:
            for row in table:
                if is_row_blank(row):
                    continue
                cells = [clean_extracted_text(c) for c in row]
                cell0 = cells[0] if cells else ""

                if is_sap_header_row(row) or is_metadata_row(row):
                    continue

                if cell0 and not _is_junk_class_candidate(cell0):
                    rest_cells_have_data = any(cells[1:])
                    if current_class and class_name_open and not rest_cells_have_data:
                        # pure wrap continuation of the class name itself
                        current_class_new = (current_class + " " + cell0).strip()
                        if current_class_new != current_class and current_class in parsed:
                            parsed[current_class_new] = parsed.pop(current_class)
                        current_class = current_class_new
                        continue
                    current_class = cell0
                    parsed.setdefault(current_class, [])
                    class_name_open = looks_incomplete(current_class)
                    if not rest_cells_have_data:
                        continue
                elif cell0:
                    class_name_open = False

                if not current_class:
                    continue

                risk = cells[1] if len(cells) > 1 else ""
                assertions = split_assertions(cells[3] if len(cells) > 3 else "") or [""]
                substantive = next(
                    (cells[idx] for idx in range(11, 4, -1) if len(cells) > idx and cells[idx]), ""
                )
                control = cells[4] if len(cells) > 4 else ""

                if not any([risk, control, substantive]) and not any(assertions):
                    continue

                recs = parsed[current_class]
                merged_any = False
                if not cell0 and recs:
                    # Empty class-name cell: this row might be a genuine
                    # continuation of the previous record's wrapped text
                    # (per-field, based on trailing-punctuation/conjunction
                    # evidence), OR a fresh additional risk entry under the
                    # same still-open class. We only fold it into the
                    # previous record for the specific field(s) that look
                    # cut off - any field without that evidence still gets
                    # a fresh record below.
                    prev = recs[-1]
                    if risk and looks_incomplete(prev["Risk"]):
                        prev["Risk"] = (prev["Risk"] + " " + risk).strip()
                        merged_any = True
                        risk = ""
                    if control and looks_incomplete(prev["Control Activity"]):
                        prev["Control Activity"] = (prev["Control Activity"] + " " + control).strip()
                        merged_any = True
                        control = ""
                    if substantive and looks_incomplete(prev["Substantive Testing Procedures"]):
                        prev["Substantive Testing Procedures"] = (
                            prev["Substantive Testing Procedures"] + " " + substantive
                        ).strip()
                        merged_any = True
                        substantive = ""
                    if merged_any:
                        for a in assertions:
                            if a and a not in prev["Assertions"]:
                                prev["Assertions"].append(a)

                if not (merged_any and not any([risk, control, substantive])):
                    if risk or control or substantive or any(assertions):
                        recs.append({
                            "Risk": risk,
                            "Assertions": assertions,
                            "Control Activity": control,
                            "Substantive Testing Procedures": substantive,
                        })

    if DEBUG_EXTRACTION:
        logger.info("sap_classes=%d", len(parsed))
    return parsed


# ===================================================================
# Row-height estimation for long wrapped text
# ===================================================================
def estimate_row_height(texts, col_widths_chars, min_h=15, max_h=409, line_h=14):
    """Approximate a row height (points) tall enough to show wrapped
    text, based on text length vs. column width. `texts` and
    `col_widths_chars` are parallel lists (one per column contributing
    to this row's height). Excel's own auto-fit is imperfect, but the
    underlying cell values are always complete regardless of visual
    height, so this is purely cosmetic.
    """
    max_lines = 1
    for text, width_chars in zip(texts, col_widths_chars):
        if not text:
            continue
        chars_per_line = max(8, int(width_chars * 1.8))
        explicit_lines = text.count("\n") + 1
        wrapped_lines = max(1, math.ceil(len(text) / chars_per_line))
        max_lines = max(max_lines, explicit_lines, wrapped_lines)
    return min(max_h, max(min_h, max_lines * line_h + 6))


# ===================================================================
# Shared openpyxl style objects (created once per workbook build and
# reused across cells instead of re-instantiating Font/Border/Fill for
# every single cell).
# ===================================================================
class Styles:
    def __init__(self):
        self.thin = Side(style="thin")
        self.medium = Side(style="medium")
        self.border_thin = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)

        self.fill_title = PatternFill("solid", fgColor="1F4E79")
        self.fill_blue = PatternFill("solid", fgColor="D9EAF7")
        self.fill_bhdr = PatternFill("solid", fgColor="2E75B6")
        self.fill_grey = PatternFill("solid", fgColor="D9D9D9")
        self.fill_yellow = PatternFill("solid", fgColor="FFD966")
        self.fill_pink = PatternFill("solid", fgColor="C55A11")
        self.fill_green = PatternFill("solid", fgColor="C6EFCE")
        self.fill_white = PatternFill("solid", fgColor="FFFFFF")
        self.fill_stripe = PatternFill("solid", fgColor="EBF3FB")

        self.font_normal = Font(size=10, name="Arial")
        self.font_bold = Font(bold=True, size=10, name="Arial")
        self.font_bold_white = Font(bold=True, size=10, name="Arial", color="FFFFFF")
        self.font_title = Font(bold=True, size=13, name="Arial", color="FFFFFF")

        self.align_left_top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
        self.align_left_center_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.align_left_center_nowrap = Alignment(horizontal="left", vertical="center", wrap_text=False)
        self.align_center_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.align_center_top_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)

    def outer_med(self, ws, r1, r2, c1, c2):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(r, c).border = Border(
                    left=self.medium if c == c1 else self.thin,
                    right=self.medium if c == c2 else self.thin,
                    top=self.medium if r == r1 else self.thin,
                    bottom=self.medium if r == r2 else self.thin,
                )

    def apply(self, cell, bold=False, fill=None, ha="left", va="center", wrap=True,
              sz=10, color="000000"):
        cell.font = Font(bold=bold, size=sz, color=color, name="Arial")
        cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
        cell.border = self.border_thin
        if fill:
            cell.fill = fill


# ===================================================================
# AWP 5.2 workbook generation
# ===================================================================
def generate_sap_workbook_bytes(tables_by_page, header_info) -> bytes:
    t0 = time.time()
    parsed_data = parse_sap_records(tables_by_page)

    wb = Workbook()
    wb.remove(wb.active)
    S = Styles()

    used_names = {}
    for full_class_name, records in parsed_data.items():
        sheet_name = make_excel_sheet_name(full_class_name, used_names)
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False

        seen_risks = {}
        for entry in records:
            rt = entry["Risk"].strip()
            if not rt:
                continue
            seen_risks.setdefault(rt, [])
            for a in entry["Assertions"]:
                if a and a not in seen_risks[rt]:
                    seen_risks[rt].append(a)

        rac = [(r, a) for r, alist in seen_risks.items() for a in (alist or [""])]
        if not rac:
            rac = [("Risk", "Relevant Assertions")]

        DYN_START = 7
        REM_COL = DYN_START + len(rac)
        LAST_COL = REM_COL

        for col, w in [("A", 2), ("B", 20), ("C", 26), ("D", 28), ("E", 32), ("F", 20), ("G", 26), ("H", 14), ("I", 14)]:
            ws.column_dimensions[col].width = w
        for ci in range(DYN_START, REM_COL):
            ws.column_dimensions[get_column_letter(ci)].width = 22
        ws.column_dimensions[get_column_letter(REM_COL)].width = 28

        ws.row_dimensions[1].height = 28
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=LAST_COL)
        # Full, untruncated class name in the title - only the sheet tab is shortened.
        ws["B1"] = f"AWP 5.2 Performing Substantive Audit Procedures - {full_class_name}"
        S.apply(ws["B1"], bold=True, sz=13, ha="center", fill=S.fill_title, color="FFFFFF")

        for r, lbl, key in [(2, "Name of the Entity :", "entity_name"), (3, "Period of Audit :", "audit_period")]:
            ws.row_dimensions[r].height = 22
            S.apply(ws.cell(r, 2, lbl), bold=True, fill=S.fill_blue, wrap=False)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=LAST_COL)
            ws.cell(r, 3, header_info[key])
            for c in range(2, LAST_COL + 1):
                ws.cell(r, c).border = S.border_thin
                ws.cell(r, c).font = S.font_normal
                if c >= 3:
                    ws.cell(r, c).alignment = S.align_left_center_nowrap
            S.apply(ws.cell(r, 2), bold=True, fill=S.fill_blue, wrap=False)
        S.outer_med(ws, 2, 3, 2, LAST_COL)

        ws.row_dimensions[4].height = 6
        for r in range(5, 9):
            ws.row_dimensions[r].height = 22

        for merge, val, ref in [
            ("B5:C5", "Prepared by", "B5"), ("D5:E5", "Signature", "D5"),
            ("F5:G5", "Reviewed & agreed by", "F5"), ("H5:I5", "Signature", "H5"),
        ]:
            ws.merge_cells(merge)
            S.apply(ws[ref], bold=True, fill=S.fill_bhdr, ha="center", color="FFFFFF", sz=10)
            ws[ref] = val

        for r, lbl, ak, rk in [
            (6, "Name", "assessed_name", "reviewed_name"),
            (7, "Designation", "assessed_designation", "reviewed_designation"),
            (8, "Date", "assessed_date", "reviewed_date"),
        ]:
            for col, key, f_ in [(2, lbl, S.fill_grey), (6, lbl, S.fill_grey)]:
                S.apply(ws.cell(r, col, key), bold=True, fill=f_, wrap=False)
            for col, key in [(3, ak), (7, rk)]:
                c = ws.cell(r, col, header_info[key])
                c.alignment = S.align_left_center_nowrap
                c.font = S.font_normal
                c.border = S.border_thin

        ws.merge_cells(start_row=6, start_column=4, end_row=8, end_column=5)
        ws.merge_cells(start_row=6, start_column=8, end_row=8, end_column=9)
        for r in range(5, 9):
            for c in range(2, 10):
                ws.cell(r, c).border = S.border_thin
        S.outer_med(ws, 5, 8, 2, 9)

        ws.row_dimensions[9].height = 6

        ws.row_dimensions[10].height = 20
        ws.merge_cells(start_row=10, start_column=2, end_row=10, end_column=LAST_COL)
        ws["B10"] = "STEP 1 : Trace risks, control activity, substantive audit procedures and relevant audit assertions"
        S.apply(ws["B10"], bold=True, sz=10, fill=S.fill_pink, ha="left", color="FFFFFF")

        ws.row_dimensions[11].height = 18
        ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=LAST_COL)
        ws["B11"] = f"Significant COTABD:  {full_class_name}"
        S.apply(ws["B11"], bold=True, sz=10, fill=S.fill_blue, ha="left")

        ws.row_dimensions[12].height = 32
        for ci, hdr in [(2, "Risk Description"), (3, "Relevant Assertions"), (4, "Control Activity"), (5, "Substantive Testing Procedures")]:
            S.apply(ws.cell(12, ci, hdr), bold=True, fill=S.fill_bhdr, ha="center", color="FFFFFF", sz=10)
        S.outer_med(ws, 12, 12, 2, 5)

        s1_start = 13
        col_widths = {2: 20, 3: 26, 4: 28, 5: 32}
        for i, entry in enumerate(records):
            rf = S.fill_stripe if i % 2 else S.fill_white
            row_idx = s1_start + i
            texts = [entry["Risk"], " / ".join(entry["Assertions"]), entry["Control Activity"], entry["Substantive Testing Procedures"]]
            ws.row_dimensions[row_idx].height = estimate_row_height(texts, [col_widths[c] for c in (2, 3, 4, 5)], min_h=48)
            for ci, val in zip([2, 3, 4, 5], texts):
                c = ws.cell(row_idx, ci, val)
                c.font = S.font_normal
                c.alignment = S.align_left_top_wrap
                c.border = S.border_thin
                c.fill = rf

        s1_end = s1_start + max(len(records) - 1, 0)
        if records:
            S.outer_med(ws, s1_start, s1_end, 2, 5)

        s2_title = s1_end + 2
        ws.row_dimensions[s2_title - 1].height = 6
        ws.row_dimensions[s2_title].height = 20
        ws.merge_cells(start_row=s2_title, start_column=2, end_row=s2_title, end_column=LAST_COL)
        ws.cell(s2_title, 2, "STEP 2 : Substantive audit procedures performed")
        S.apply(ws.cell(s2_title, 2), bold=True, fill=S.fill_pink, ha="left", color="FFFFFF", sz=10)

        h1, h2 = s2_title + 1, s2_title + 2
        ws.row_dimensions[h1].height = ws.row_dimensions[h2].height = 36

        for ci, hdr in [(2, "Sl\nNo"), (3, "Date"), (4, "Voucher\nNo."), (5, "Voucher\nAmount (Nu.)"), (6, "Details"), (REM_COL, "Remarks")]:
            ws.merge_cells(start_row=h1, start_column=ci, end_row=h2, end_column=ci)
            S.apply(ws.cell(h1, ci, hdr), bold=True, fill=S.fill_bhdr, ha="center", color="FFFFFF", sz=10)

        risk_groups = {}
        for col_i, (rt, _) in enumerate(rac, DYN_START):
            risk_groups.setdefault(rt, []).append(col_i)

        for rt, cols in risk_groups.items():
            col_s, col_e = cols[0], cols[-1]
            if col_s != col_e:
                ws.merge_cells(start_row=h1, start_column=col_s, end_row=h1, end_column=col_e)
            c = ws.cell(h1, col_s, rt)
            c.font = S.font_bold
            c.alignment = S.align_center_center_wrap
            c.fill = S.fill_yellow
            for ci2 in cols:
                ws.cell(h1, ci2).border = S.border_thin

        for col_i, (_, asrt) in enumerate(rac, DYN_START):
            c = ws.cell(h2, col_i, asrt)
            c.font = S.font_bold
            c.alignment = S.align_center_center_wrap
            c.fill = S.fill_yellow
            c.border = S.border_thin

        dr_start, dr_end = h2 + 1, h2 + 20
        ynv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        ws.add_data_validation(ynv)

        for i, r in enumerate(range(dr_start, dr_end + 1)):
            rf = S.fill_stripe if i % 2 else S.fill_white
            ws.row_dimensions[r].height = 18
            for c in range(2, LAST_COL + 1):
                cell = ws.cell(r, c)
                cell.border = S.border_thin
                cell.fill = rf
                cell.alignment = S.align_left_center_wrap
            ws.cell(r, 2, i + 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(r, 2).font = S.font_normal
            ws.cell(r, 2).fill = rf
            for c in range(DYN_START, REM_COL):
                ynv.add(ws.cell(r, c))

        S.outer_med(ws, s2_title, dr_end, 2, LAST_COL)

        conc = dr_end + 3
        ws.row_dimensions[dr_end + 2].height = 6
        ws.row_dimensions[conc].height = 60
        c = ws.cell(conc, 2, "Overall Conclusion:")
        c.font = S.font_bold
        c.alignment = S.align_left_center_wrap
        c.fill = S.fill_green
        c.border = S.border_thin
        ws.merge_cells(start_row=conc, start_column=3, end_row=conc, end_column=LAST_COL)
        ws.cell(conc, 3).alignment = S.align_left_top_wrap
        ws.cell(conc, 3).fill = S.fill_white
        for ci in range(3, LAST_COL + 1):
            ws.cell(conc, ci).border = S.border_thin
        S.outer_med(ws, conc, conc, 2, LAST_COL)

    if len(wb.sheetnames) == 0:
        wb.create_sheet("No Data")

    out = BytesIO()
    wb.save(out)
    logger.info("sap_workbook_built sheets=%d duration=%.2fs", len(wb.sheetnames), time.time() - t0)
    return out.getvalue()


# ===================================================================
# AWP 5.8 (MAP / Table C) workbook generation
# ===================================================================
def generate_map_workbook_bytes(tables_by_page, header_info) -> bytes:
    t0 = time.time()
    map_rows, table_c_label = extract_table_c_rows(tables_by_page)

    wb = Workbook()
    wb.remove(wb.active)
    S = Styles()

    used_names = {}
    for full_class_name, procedure_text in map_rows:
        sheet_name = make_excel_sheet_name(full_class_name, used_names)
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False

        LAST_COL = 9  # B..I, matches template

        for col, w in [("A", 2), ("B", 20), ("C", 26), ("D", 28), ("E", 32), ("F", 20), ("G", 22), ("H", 14), ("I", 28)]:
            ws.column_dimensions[col].width = w

        ws.row_dimensions[1].height = 28
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=LAST_COL)
        ws["B1"] = f"AWP 5.8 Performing Minimum Audit Procedures - {full_class_name}"
        S.apply(ws["B1"], bold=True, sz=13, ha="center", fill=S.fill_title, color="FFFFFF")

        for r, lbl, key in [(2, "Name of the Entity :", "entity_name"), (3, "Period of Audit :", "audit_period")]:
            ws.row_dimensions[r].height = 22
            S.apply(ws.cell(r, 2, lbl), bold=True, fill=S.fill_blue, wrap=False)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=LAST_COL)
            ws.cell(r, 3, header_info[key])
            for c in range(2, LAST_COL + 1):
                ws.cell(r, c).border = S.border_thin
                ws.cell(r, c).font = S.font_normal
                if c >= 3:
                    ws.cell(r, c).alignment = S.align_left_center_nowrap
            S.apply(ws.cell(r, 2), bold=True, fill=S.fill_blue, wrap=False)
        S.outer_med(ws, 2, 3, 2, LAST_COL)

        ws.row_dimensions[4].height = 6
        for r in range(5, 9):
            ws.row_dimensions[r].height = 22

        for merge, val, ref in [
            ("B5:C5", "Prepared by", "B5"), ("D5:E5", "Signature", "D5"),
            ("F5:G5", "Reviewed & agreed by", "F5"), ("H5:I5", "Signature", "H5"),
        ]:
            ws.merge_cells(merge)
            S.apply(ws[ref], bold=True, fill=S.fill_bhdr, ha="center", color="FFFFFF", sz=10)
            ws[ref] = val

        for r, lbl, ak, rk in [
            (6, "Name", "assessed_name", "reviewed_name"),
            (7, "Designation", "assessed_designation", "reviewed_designation"),
            (8, "Date", "assessed_date", "reviewed_date"),
        ]:
            for col, key, f_ in [(2, lbl, S.fill_grey), (6, lbl, S.fill_grey)]:
                S.apply(ws.cell(r, col, key), bold=True, fill=f_, wrap=False)
            for col, key in [(3, ak), (7, rk)]:
                c = ws.cell(r, col, header_info[key])
                c.alignment = S.align_left_center_nowrap
                c.font = S.font_normal
                c.border = S.border_thin

        ws.merge_cells(start_row=6, start_column=4, end_row=8, end_column=5)
        ws.merge_cells(start_row=6, start_column=8, end_row=8, end_column=9)
        for r in range(5, 9):
            for c in range(2, 10):
                ws.cell(r, c).border = S.border_thin
        S.outer_med(ws, 5, 8, 2, 9)

        ws.row_dimensions[9].height = 6

        ws.row_dimensions[10].height = 20
        ws.merge_cells(start_row=10, start_column=2, end_row=10, end_column=LAST_COL)
        ws["B10"] = "STEP 1 : Trace minimum audit procedure"
        S.apply(ws["B10"], bold=True, sz=10, fill=S.fill_pink, ha="left", color="FFFFFF")

        ws.row_dimensions[11].height = 17
        ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=LAST_COL)
        ws["B11"] = f"{table_c_label}:  {full_class_name}"
        S.apply(ws["B11"], bold=True, sz=10, fill=S.fill_blue, ha="left")

        ws.row_dimensions[12].height = 17
        ws.merge_cells(start_row=12, start_column=2, end_row=12, end_column=5)
        S.apply(ws.cell(12, 2, "Substantive Testing Procedures"), bold=True, fill=S.fill_bhdr, ha="center", color="FFFFFF", sz=10)

        row13_h = estimate_row_height([procedure_text], [26], min_h=17)
        ws.row_dimensions[13].height = row13_h
        ws.merge_cells(start_row=13, start_column=2, end_row=13, end_column=5)
        c = ws.cell(13, 2, procedure_text)
        c.font = S.font_normal
        c.alignment = S.align_center_top_wrap
        c.fill = S.fill_stripe
        for ci in range(2, 6):
            ws.cell(13, ci).border = S.border_thin
        S.outer_med(ws, 12, 13, 2, 5)

        ws.row_dimensions[14].height = 17

        s2_title = 15
        ws.row_dimensions[s2_title].height = 20
        ws.merge_cells(start_row=s2_title, start_column=2, end_row=s2_title, end_column=7)
        ws.cell(s2_title, 2, "STEP 2 : Minimum audit procedures performed")
        S.apply(ws.cell(s2_title, 2), bold=True, fill=S.fill_pink, ha="left", color="FFFFFF", sz=10)

        h1, h2 = s2_title + 1, s2_title + 2
        ws.row_dimensions[h1].height = ws.row_dimensions[h2].height = 36

        for ci, hdr in [(2, "Sl\nNo"), (3, "Date"), (4, "Voucher\nNo."), (5, "Voucher\nAmount (Nu.)"), (6, "Details"), (7, "Remarks")]:
            ws.merge_cells(start_row=h1, start_column=ci, end_row=h2, end_column=ci)
            S.apply(ws.cell(h1, ci, hdr), bold=True, fill=S.fill_bhdr, ha="center", color="FFFFFF", sz=10)

        dr_start, dr_end = h2 + 1, h2 + 20
        for i, r in enumerate(range(dr_start, dr_end + 1)):
            rf = S.fill_stripe if i % 2 else S.fill_white
            ws.row_dimensions[r].height = 18
            for c in range(2, 8):
                cell = ws.cell(r, c)
                cell.border = S.border_thin
                cell.fill = rf
                cell.alignment = S.align_left_center_wrap
            ws.cell(r, 2, i + 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(r, 2).font = S.font_normal
            ws.cell(r, 2).fill = rf

        S.outer_med(ws, s2_title, dr_end, 2, 7)

        conc = dr_end + 3
        ws.row_dimensions[dr_end + 2].height = 6
        ws.row_dimensions[conc].height = 17
        c = ws.cell(conc, 2, "Overall Conclusion:")
        c.font = S.font_bold
        c.alignment = S.align_left_center_wrap
        c.fill = S.fill_green
        c.border = S.border_thin
        ws.merge_cells(start_row=conc, start_column=3, end_row=conc, end_column=LAST_COL)
        ws.cell(conc, 3).alignment = S.align_left_top_wrap
        ws.cell(conc, 3).fill = S.fill_white
        for ci in range(3, LAST_COL + 1):
            ws.cell(conc, ci).border = S.border_thin
        S.outer_med(ws, conc, conc, 2, LAST_COL)

    if len(wb.sheetnames) == 0:
        wb.create_sheet("No Data")

    out = BytesIO()
    wb.save(out)
    logger.info("map_workbook_built sheets=%d duration=%.2fs", len(wb.sheetnames), time.time() - t0)
    return out.getvalue()


# ===================================================================
# @st.cache_data wrappers
#
# All three cached functions take only cacheable inputs (bytes, or the
# plain lists/dicts produced by extract_pdf_tables) and return either a
# dataclass of plain data or raw `bytes` - never a Streamlit
# UploadedFile, an open pdfplumber object, or a mutable BytesIO stream.
# The same pdf_bytes will always hash to the same cache key, so
# re-running "Extract to Excel" on an already-seen PDF, or two users
# uploading the same report, reuse the cached result instead of
# re-parsing/re-building workbooks.
# ===================================================================
@st.cache_data(show_spinner=False, max_entries=64)
def cached_extract_pdf_tables(pdf_bytes: bytes):
    result = extract_pdf_tables(pdf_bytes)
    # Return plain (picklable, hashable-by-value) data rather than the
    # dataclass instance, keeping the cache boundary strictly primitive.
    return {
        "tables_by_page": result.tables_by_page,
        "header_info": result.header_info,
        "num_pages": result.num_pages,
        "num_tables": result.num_tables,
        "pdf_hash_prefix": result.pdf_hash_prefix,
    }


@st.cache_data(show_spinner=False, max_entries=64)
def cached_generate_sap_workbook(tables_by_page, header_info) -> bytes:
    return generate_sap_workbook_bytes(tables_by_page, header_info)


@st.cache_data(show_spinner=False, max_entries=64)
def cached_generate_map_workbook(tables_by_page, header_info) -> bytes:
    return generate_map_workbook_bytes(tables_by_page, header_info)


# ===================================================================
# Streamlit page config + CSS (preserved from the original design)
# ===================================================================
def _configure_page():
    st.set_page_config(page_title="AWP 5.2 & 5.8 Automation", layout="centered", initial_sidebar_state="collapsed")
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
* { font-family: 'DM Sans', sans-serif; }

[data-testid="stAppViewContainer"] { background: #f8fafc; min-height: 100vh; }
[data-testid="stHeader"] { background: transparent; }
.block-container { max-width: 720px; padding-top: 0rem !important; padding-bottom: 2rem !important; }
#MainMenu, footer { visibility: hidden; }

/* Hero */
.hero { text-align: center; padding: 0.2rem 0 0.5rem; }
.hero .badge { display: inline-flex; align-items: center; gap: 6px; background: #dcfce7; color: #15803d; border: 1px solid #bbf7d0; border-radius: 999px; padding: 5px 14px; font-size: 12px; font-weight: 600; margin-bottom: 6px; }
.hero .badge-dot { width: 7px; height: 7px; border-radius: 50%; background: #22c55e; animation: pulse 2s infinite; }
[data-testid="stImage"] { margin: -8px 0 -6px 0 !important; }
@keyframes pulse { 0%,100%{opacity:1;transform:scale(1)} 50%{opacity:.5;transform:scale(1.35)} }
.hero h1 { font-size: 34px; font-weight: 700; line-height: 1.18; letter-spacing: -.04em; color: #0f172a; margin: 0 0 0.5px 0; }
.hero h1 span.green { color: #16a34a; }
.hero p { font-size: 15px; line-height: 1.7; color: #64748b; margin: 0; }

/* Steps bar */
.steps-bar { display: flex; align-items: center; justify-content: center; background: #fff; border: 1px solid #e2e8f0; border-radius: 18px; padding: 12px 20px; margin: 1rem 0; box-shadow: 0 1px 2px rgba(0,0,0,0.05); }
.step-item { display: flex; align-items: center; justify-content: center; gap: 10px; flex: 1; }
.step-num { width: 32px; height: 32px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 13px; font-weight: 700; border: 2px solid #dbe4ee; background: #f8fafc; color: #64748b; }
.step-num.done   { background: #16a34a; border-color: #16a34a; color: white; }
.step-num.active { background: linear-gradient(135deg,#16a34a,#15803d); border-color: #15803d; color: white; }
.step-text { font-size: 13px; font-weight: 500; color: #64748b; }
.step-text.active { color: #0f172a; }
.step-text.done   { color: #15803d; }
.step-connector      { width: 44px; height: 2px; background: #e2e8f0; border-radius: 999px; }
.step-connector.done { background: #22c55e; }

[data-testid="stFileUploaderDropzone"] {
    background: #ffffff !important;
    border: 2px dashed #cbd5e1 !important;
    border-radius: 20px !important;
    padding: 1.5rem !important;
    width: 320px !important;
    height: 200px !important;
    margin: 0 auto !important;
    box-shadow: 0 6px 18px rgba(15, 23, 42, 0.05) !important;
    transition: border-color 0.2s ease, box-shadow 0.2s ease !important;

    display: flex !important;
    flex-direction: column !important;
    align-items: center !important;
    justify-content: center !important;
    gap: 14px !important;
    text-align: center !important;
}

[data-testid="stFileUploaderDropzone"]:hover {
    border-color: #94a3b8 !important;
    background: #ffffff !important;
    box-shadow: 0 8px 22px rgba(15, 23, 42, 0.07) !important;
    transform: none !important;
}

/* Browse files button */
[data-testid="stFileUploaderDropzone"] button {
    order: -1 !important;
    background: linear-gradient(135deg, #16a34a, #15803d) !important;
    border: none !important;
    border-radius: 999px !important;
    min-width: 210px !important;
    height: 50px !important;
    padding: 0 24px !important;
    margin: 0 auto 6px auto !important;
    box-shadow: none !important;

    display: flex !important;
    align-items: center !important;
    justify-content: center !important;

    font-size: 0 !important;
    color: transparent !important;
}

[data-testid="stFileUploaderDropzone"] button::after {
    content: "Browse files";
    font-size: 15px !important;
    line-height: 1 !important;
    font-weight: 700 !important;
    color: #ffffff !important;
}

[data-testid="stFileUploaderDropzone"] button:hover {
    background: linear-gradient(135deg, #22c55e, #16a34a) !important;
    transform: translateY(-1px) !important;
}

[data-testid="stFileUploaderDropzone"] button *,
[data-testid="stFileUploaderDropzone"] button p,
[data-testid="stFileUploaderDropzone"] button span {
    display: none !important;
}

/* Instruction area */
[data-testid="stFileUploaderDropzoneInstructions"] {
    display: flex !important;
    flex-direction: column !important;
    align-items: center !important;
    justify-content: center !important;
    text-align: center !important;
    gap: 4px !important;
    margin: 0 !important;
}

/* Upload icon */
[data-testid="stFileUploaderDropzoneInstructions"] > span,
[data-testid="stFileUploaderDropzoneInstructions"] svg {
    display: block !important;
    color: #94a3b8 !important;
    fill: #94a3b8 !important;
    width: 34px !important;
    height: 34px !important;
    margin: 0 auto 4px auto !important;
}

/* Hide Streamlit wording */
[data-testid="stFileUploaderDropzoneInstructions"] div[data-testid="stMarkdownContainer"],
[data-testid="stFileUploaderDropzoneInstructions"] > div:first-child > span {
    display: none !important;
}

[data-testid="stFileUploaderDropzoneInstructions"]::before {
    content: "Drag and drop file here";
    display: block !important;
    font-size: 15px !important;
    line-height: 1.2 !important;
    font-weight: 500 !important;
    color: #374151 !important;
}

[data-testid="stFileUploaderDropzoneInstructions"]::after {
    content: "Limit 200MB per file • PDF";
    display: block !important;
    font-size: 12px !important;
    line-height: 1.2 !important;
    font-weight: 400 !important;
    color: #9ca3af !important;
}

/* Button centering wrapper */
.btn-center { display: flex; justify-content: center; margin: 0.5rem 0; }

/* All stButtons — green, fixed width */
.stButton > button {
    background: linear-gradient(135deg, #16a34a, #15803d) !important;
    border: none !important; border-radius: 40px !important;
    font-weight: 600 !important; color: white !important;
    transition: all 0.25s ease !important;
    width: auto !important; min-width: 200px; max-width: 280px;
    padding: 10px 32px !important;
    box-shadow: 0 4px 12px rgba(22,163,74,0.3);
    display: block; margin: 0 auto;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #22c55e, #16a34a) !important;
    transform: translateY(-2px);
    box-shadow: 0 10px 20px -5px rgba(22,163,74,0.5) !important;
    color: white !important;
}

/* Remove / Start Over button — muted style */
.remove-btn .stButton > button {
    background: #f1f5f9 !important;
    color: #64748b !important;
    border: 1px solid #e2e8f0 !important;
    box-shadow: none !important;
    font-weight: 500 !important;
    font-size: 13px !important;
    min-width: 120px; max-width: 160px;
    padding: 8px 20px !important;
}
.remove-btn .stButton > button:hover {
    background: #fee2e2 !important;
    color: #dc2626 !important;
    border-color: #fecaca !important;
    box-shadow: none !important;
    transform: translateY(-1px);
}

/* Download button — green, fixed width */
[data-testid="stDownloadButton"] button {
    background: linear-gradient(135deg, #16a34a, #15803d) !important;
    border: none !important; border-radius: 40px !important;
    font-weight: 600 !important; transition: all 0.25s ease !important;
    width: auto !important; min-width: 220px; max-width: 300px;
    padding: 10px 32px !important;
    box-shadow: 0 4px 12px rgba(22,163,74,0.3);
    display: block; margin: 0 auto;
}
[data-testid="stDownloadButton"] button:hover {
    background: linear-gradient(135deg, #22c55e, #16a34a) !important;
    transform: translateY(-2px);
    box-shadow: 0 10px 20px -5px rgba(22,163,74,0.5) !important;
}

/* Input */
[data-testid="stTextInput"] input { border-radius: 40px; border: 1px solid #e2e8f0; padding: 10px 16px; font-size: 14px; }

/* Upload success */
.upload-success { background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 20px; padding: 1rem 1.2rem; text-align: center; margin: 0.5rem 0; animation: fadeInUp 0.3s ease; }
.upload-success .filename { font-weight: 600; color: #166534; word-break: break-all; }

/* Result box */
.result-box { background: linear-gradient(135deg, #f0fdf4, #dcfce7); border-radius: 24px; padding: 1.5rem; margin-top: 1.5rem; margin-bottom: 1.75rem; border: 1px solid #bbf7d0; text-align: center; animation: fadeInUp 0.4s ease; }
.result-title { font-weight: 700; font-size: 18px; color: #166534; margin-bottom: 1rem; }
.output-card { background: #ffffff; border: 1px solid #bbf7d0; border-bottom: none; border-radius: 16px 16px 0 0; padding: 1.1rem 1.1rem 0.9rem; margin-bottom: 0; text-align: left; }
.output-card-title { font-weight: 700; font-size: 16px; color: #0f172a; letter-spacing: -0.01em; }
.output-card-desc { font-size: 12.5px; color: #64748b; margin-top: 4px; line-height: 1.4; }
[data-testid="stDownloadButton"] { margin-top: -6px; }
[data-testid="stDownloadButton"] button {
    border-radius: 0 0 16px 16px !important;
    width: 100% !important; max-width: none !important; min-width: 0 !important;
    margin: 0 !important;
    box-shadow: none !important;
    color: #ffffff !important;
}
[data-testid="stDownloadButton"] button p { color: #ffffff !important; }
.output-actions { margin-top: 2rem; padding-top: 1.25rem; border-top: 1px solid #e2e8f0; }
@keyframes fadeInUp { from{opacity:0;transform:translateY(15px)} to{opacity:1;transform:translateY(0)} }

/* Error box */
.error-box { background: #fef2f2; border: 1px solid #fecaca; border-radius: 20px; padding: 1rem 1.2rem; text-align: center; margin: 0.5rem 0; color: #b91c1c; }

/* Footer */
.app-footer { text-align: center; font-size: 12px; color: #94a3b8; border-top: 1px solid #e2e8f0; padding-top: 1.5rem; margin-top: 2rem; }

@media (max-width: 640px) {
    .hero h1 { font-size: 24px; }
    .steps-bar { flex-wrap: wrap; gap: 12px; }
    .step-connector { display: none; }
}
</style>
""", unsafe_allow_html=True)


def steps_bar(current: int):
    labels, icons = ["Upload", "Automate", "Download"], ["☁", "⚙", "⬇"]
    html = '<div class="steps-bar">'
    for i, (lbl, icon) in enumerate(zip(labels, icons), 1):
        nc = "done" if i < current else ("active" if i == current else "")
        dot = "✓" if i < current else icon
        html += f'<div class="step-item"><div class="step-num {nc}">{dot}</div><span class="step-text {nc}">{lbl}</span></div>'
        if i < len(labels):
            html += f'<div class="step-connector {"done" if i < current else ""}"></div>'
    st.markdown(html + '</div>', unsafe_allow_html=True)


def _reset_session_state():
    for key in ("pdf_bytes", "pdf_hash", "uploaded_file_name", "sap_output",
                "map_output", "excel_ready", "processing_error",
                "elapsed_seconds", "extraction_meta"):
        st.session_state[key] = None
    st.session_state["excel_ready"] = False


def _init_session_state():
    defaults = {
        "pdf_bytes": None, "pdf_hash": None, "uploaded_file_name": "",
        "sap_output": None, "map_output": None, "excel_ready": False,
        "processing_error": None, "elapsed_seconds": None, "extraction_meta": None,
    }
    for key, default in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = default


def _process_pdf(pdf_bytes: bytes, filename: str):
    """Runs the full extraction + both workbook builds, with progress
    messages, timing, and exception handling. Never leaves the UI stuck
    in a processing state - any failure is captured and surfaced to the
    user as a friendly message instead of a crash."""
    t_start = time.time()
    try:
        with st.status("Processing your PDF…", expanded=True) as status:
            st.write("Validating PDF…")
            pdf_hash_prefix = compute_pdf_hash(pdf_bytes)[:12]
            logger.info("processing_start pdf_hash=%s filename=%s", pdf_hash_prefix, filename)

            st.write("Reading pages and extracting tables…")
            t_extract = time.time()
            extraction = cached_extract_pdf_tables(pdf_bytes)
            extract_duration = time.time() - t_extract

            st.write(
                f"Found {extraction['num_pages']} page(s) and "
                f"{extraction['num_tables']} table(s). Building AWP 5.2 workbook…"
            )
            t_sap = time.time()
            sap_bytes = cached_generate_sap_workbook(extraction["tables_by_page"], extraction["header_info"])
            sap_duration = time.time() - t_sap

            st.write("Building AWP 5.8 workbook…")
            t_map = time.time()
            map_bytes = cached_generate_map_workbook(extraction["tables_by_page"], extraction["header_info"])
            map_duration = time.time() - t_map

            st.write("Finalizing files…")
            status.update(label="Done", state="complete", expanded=False)

        elapsed = time.time() - t_start
        st.session_state.sap_output = sap_bytes
        st.session_state.map_output = map_bytes
        st.session_state.excel_ready = True
        st.session_state.elapsed_seconds = elapsed
        st.session_state.processing_error = None
        st.session_state.extraction_meta = {
            "num_pages": extraction["num_pages"],
            "num_tables": extraction["num_tables"],
        }

        logger.info(
            "processing_done pdf_hash=%s pages=%d tables=%d "
            "extract_s=%.2f sap_s=%.2f map_s=%.2f total_s=%.2f",
            pdf_hash_prefix, extraction["num_pages"], extraction["num_tables"],
            extract_duration, sap_duration, map_duration, elapsed,
        )

    except PDFValidationError as exc:
        logger.warning("pdf_validation_failed filename=%s reason=%s", filename, exc)
        st.session_state.processing_error = str(exc)
        st.session_state.excel_ready = False
    except Exception:
        logger.exception("processing_failed filename=%s", filename)
        st.session_state.processing_error = (
            "Something went wrong while processing this PDF. Please try again, "
            "or use a different file. If the problem continues, contact support "
            f"with reference code {compute_pdf_hash(pdf_bytes)[:8]}."
        )
        st.session_state.excel_ready = False


def tool_page():
    _configure_page()
    _init_session_state()

    logo_path = "AWP Logo.png"

    st.markdown("""
    <div class="hero">
        <div class="badge"><div class="badge-dot"></div>Automation at Work</div>
    </div>
    """, unsafe_allow_html=True)

    if os.path.exists(logo_path):
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.image(logo_path, use_container_width=True)
    else:
        st.info("Logo not found. Place 'AWP Logo.png' in the same directory.")

    st.markdown("""
    <div class="hero">
        <h1><span class="green">Smarter.</span><span class="green"> Faster.</span><span class="green"> Better.</span></h1>
        <p>Upload AWP 4.6 PDF and let automation do the rest.</p>
    </div>
    """, unsafe_allow_html=True)

    # ── STEP 1 : Upload ────────────────────────────────────────────
    if st.session_state.pdf_bytes is None:
        steps_bar(1)
        col1, col2, col3 = st.columns([0.35, 3.3, 0.35])
        with col2:
            uploaded_pdf = st.file_uploader("Choose PDF File", type=["pdf"], label_visibility="collapsed")
            if uploaded_pdf is not None:
                # Read bytes ONCE and discard the UploadedFile wrapper -
                # everything downstream works with immutable bytes only.
                pdf_bytes = uploaded_pdf.getvalue()
                st.session_state.pdf_bytes = pdf_bytes
                st.session_state.pdf_hash = compute_pdf_hash(pdf_bytes)
                st.session_state.uploaded_file_name = uploaded_pdf.name
                st.session_state.excel_ready = False
                st.session_state.sap_output = None
                st.session_state.map_output = None
                st.session_state.processing_error = None
                st.rerun()

    # ── STEP 2 : Extract ───────────────────────────────────────────
    elif not st.session_state.excel_ready:
        steps_bar(2)

        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown(f"""
            <div class="upload-success">
                ✅ <strong>PDF uploaded successfully</strong><br>
                <span class="filename">{st.session_state.uploaded_file_name}</span>
            </div>
            """, unsafe_allow_html=True)
            st.markdown('<div class="remove-btn">', unsafe_allow_html=True)
            if st.button("Remove PDF", key="remove_pdf"):
                _reset_session_state()
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

            if st.session_state.processing_error:
                st.markdown(
                    f'<div class="error-box">⚠️ {st.session_state.processing_error}</div>',
                    unsafe_allow_html=True,
                )

        if st.button("Extract to Excel"):
            _process_pdf(st.session_state.pdf_bytes, st.session_state.uploaded_file_name)
            st.rerun()

    # ── STEP 3 : Download ──────────────────────────────────────────
    else:
        steps_bar(3)

        uploaded_name = st.session_state.uploaded_file_name or "Audit_Workbook"
        base_name = os.path.splitext(uploaded_name)[0]
        safe_name = re.sub(r'[\\/*?:"<>|]', '_', base_name).strip() or "Audit_Workbook"

        elapsed = st.session_state.elapsed_seconds
        elapsed_txt = f" (generated in {elapsed:.1f}s)" if elapsed else ""
        st.markdown(
            f'<div class="result-box"><div class="result-title">'
            f'Excel generated successfully{elapsed_txt}</div>',
            unsafe_allow_html=True,
        )

        col_sap, col_map = st.columns(2)
        with col_sap:
            st.markdown('<div class="output-card"><div class="output-card-title">AWP 5.2</div><div class="output-card-desc">Substantive Audit Procedures</div></div>', unsafe_allow_html=True)
            st.download_button(
                label="Download",
                data=st.session_state.sap_output,
                file_name=f"{safe_name} AWP 5.2.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_sap",
            )
        with col_map:
            st.markdown('<div class="output-card"><div class="output-card-title">AWP 5.8</div><div class="output-card-desc">Minimum Audit Procedures</div></div>', unsafe_allow_html=True)
            st.download_button(
                label="Download",
                data=st.session_state.map_output,
                file_name=f"{safe_name} AWP 5.8.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_map",
            )

        st.markdown('<div class="output-actions">', unsafe_allow_html=True)
        st.markdown('<div class="remove-btn">', unsafe_allow_html=True)
        if st.button("Process Another PDF"):
            _reset_session_state()
            st.rerun()
        st.markdown('</div></div></div>', unsafe_allow_html=True)

    st.markdown('<div class="app-footer">© Royal Audit Authority · Supreme Audit Institution of Bhutan</div>', unsafe_allow_html=True)


if __name__ == "__main__":
    tool_page()
